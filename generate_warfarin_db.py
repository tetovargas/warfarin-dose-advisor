import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta

np.random.seed(42)
random.seed(42)

N_PATIENTS = 250
N_WEEKS = 20
START_DATE = datetime(2024, 1, 8)

# ── Genotype distributions ────────────────────────────────────────────────────
cyp2c9_genotypes = ['*1/*1', '*1/*2', '*1/*3', '*2/*2', '*2/*3', '*3/*3']
cyp2c9_weights   = [0.650,   0.200,   0.100,   0.030,   0.015,   0.005]
cyp2c9_sens      = {'*1/*1': 1.0, '*1/*2': 1.3, '*1/*3': 1.6,
                    '*2/*2': 1.7, '*2/*3': 2.1, '*3/*3': 2.8}

vkorc1_genotypes = ['GG',  'AG',  'AA']
vkorc1_weights   = [0.37,  0.49,  0.14]
vkorc1_sens      = {'GG': 0.8, 'AG': 1.0, 'AA': 1.4}

valve_brands = ['St. Jude Medical', 'Medtronic Hall', 'Carbomedics', 'On-X', 'ATS Medical']
comorbidity_pool = [
    'Atrial Fibrillation', 'Heart Failure', 'Hypertension',
    'Diabetes Mellitus', 'CKD Stage 2', 'CKD Stage 3',
    'Hypothyroidism', 'Hyperthyroidism', 'Mild Liver Disease', 'COPD',
    'Coronary Artery Disease'
]

# ── Generate patient demographics ─────────────────────────────────────────────
patients = []
for pid in range(1, N_PATIENTS + 1):
    sex    = random.choice(['M', 'F'])
    age    = int(np.clip(np.random.normal(62, 12), 25, 85))
    weight = round(np.clip(np.random.normal(78 if sex == 'M' else 68, 12), 45, 140), 1)
    height = round(np.clip(np.random.normal(175 if sex == 'M' else 162, 8), 148, 200), 1)
    bmi    = round(weight / (height / 100) ** 2, 1)

    cyp = random.choices(cyp2c9_genotypes, weights=cyp2c9_weights)[0]
    vkr = random.choices(vkorc1_genotypes, weights=vkorc1_weights)[0]

    sens = cyp2c9_sens[cyp] * vkorc1_sens[vkr]
    sens *= 1 + (age - 60) * 0.005
    sens *= 1 - (weight - 70) * 0.003
    sens  = float(np.clip(sens, 0.3, 4.0))

    n_co = random.choices([0, 1, 2, 3], weights=[0.20, 0.45, 0.25, 0.10])[0]
    comorbidities = random.sample(comorbidity_pool, n_co)

    if 'Hypothyroidism'  in comorbidities: sens *= 0.85
    if 'Hyperthyroidism' in comorbidities: sens *= 1.15
    if 'Mild Liver Disease' in comorbidities: sens *= 1.20
    if 'Heart Failure'   in comorbidities: sens *= 1.10

    # Starting dose: target INR ~3.0, INR ≈ dose * sens * 0.25
    dose_factor   = 0.25
    start_dose    = round(3.0 / (sens * dose_factor) * 2) / 2
    start_dose    = float(np.clip(start_dose, 1.0, 15.0))
    implant_date  = START_DATE - timedelta(days=random.randint(180, 3650))

    patients.append({
        'Patient_ID': f'P{pid:03d}',
        'Age': age, 'Sex': sex,
        'Weight_kg': weight, 'Height_cm': height, 'BMI': bmi,
        'Valve_Brand': random.choice(valve_brands),
        'Valve_Position': 'Mitral',
        'Implant_Date': implant_date.strftime('%Y-%m-%d'),
        'CYP2C9_Genotype': cyp,
        'VKORC1_Genotype': vkr,
        'Comorbidities': '; '.join(comorbidities) if comorbidities else 'None',
        'Target_INR_Min': 2.5, 'Target_INR_Max': 3.5,
        'Baseline_Warfarin_Dose_mg_day': start_dose,
        '_sens': sens
    })

patients_df = pd.DataFrame(patients)
print(f"Patients generated: {len(patients_df)}")

# ── Drug options (non-chronic) ────────────────────────────────────────────────
DRUGS = [
    ('Amoxicillin',      'Antibiotic',        0.40),
    ('Ciprofloxacin',    'Antibiotic',        0.80),
    ('Metronidazole',    'Antibiotic',        1.20),
    ('Trimethoprim-SMX', 'Antibiotic',        0.70),
    ('Fluconazole',      'Antifungal',        1.80),
    ('Itraconazole',     'Antifungal',        1.20),
    ('Ibuprofen',        'NSAID',             0.30),
    ('Naproxen',         'NSAID',             0.25),
    ('Omeprazole',       'PPI',               0.20),
    ('Simvastatin',      'Statin',            0.15),
    ('Levothyroxine',    'Thyroid Hormone',   0.50),
    ('Rifampin',         'Antibiotic',       -1.50),
    ('Carbamazepine',    'Anticonvulsant',   -1.20),
    ('Aspirin 81mg',     'Antiplatelet',      0.10),
]

VIT_K_FOODS = ['Spinach','Kale','Broccoli','Brussels Sprouts','Collard Greens','Cabbage','Lettuce']
VIT_K_DELTA = {'Low': 0.15, 'Normal': 0.0, 'High': -0.65}

# ── Simulate weekly records ───────────────────────────────────────────────────
weekly_records = []

for _, pt in patients_df.iterrows():
    sens   = pt['_sens']
    df     = 0.25           # dose_factor
    dose   = pt['Baseline_Warfarin_Dose_mg_day']

    # Assign chronic amiodarone to some AF patients
    chronic_amio = ('Atrial Fibrillation' in pt['Comorbidities'] and random.random() < 0.40)

    for wk in range(1, N_WEEKS + 1):
        visit_dt = START_DATE + timedelta(weeks=wk - 1)

        # ── Random events ──────────────────────────────────────────────────
        vitk    = random.choices(['Low','Normal','High'], weights=[0.20, 0.55, 0.25])[0]
        vitk_f  = ', '.join(random.sample(VIT_K_FOODS, random.randint(2,4))) if vitk=='High' else (
                  random.choice(VIT_K_FOODS) if (vitk=='Normal' and random.random()<0.5) else 'None')

        alc     = float(np.clip(
            random.choices([0, random.randint(1,7), random.randint(8,14), random.randint(15,28)],
                           weights=[0.30, 0.40, 0.20, 0.10])[0], 0, 28))
        cranb   = random.random() < 0.08
        grape   = random.random() < 0.05
        garlic  = random.random() < 0.06

        missed  = random.choices([0,1,2,3], weights=[0.70,0.18,0.08,0.04])[0]
        illness = random.choices(['None','Mild','Moderate','Severe'], weights=[0.75,0.15,0.07,0.03])[0]
        diarhea = random.random() < 0.06
        exercise= random.choices(['Sedentary','Moderate','Active'], weights=[0.30,0.50,0.20])[0]

        # Drug this week
        if chronic_amio:
            drug_name, drug_cat, drug_eff = 'Amiodarone', 'Antiarrhythmic', 1.00
        elif random.random() < 0.12:
            drug_name, drug_cat, drug_eff = random.choice(DRUGS)
        else:
            drug_name, drug_cat, drug_eff = 'None', 'None', 0.0

        # ── INR calculation ────────────────────────────────────────────────
        base  = dose * sens * df
        delta = 0.0
        delta += VIT_K_DELTA[vitk]
        delta += min(alc * 0.05, 0.8)
        delta += 0.25 if cranb  else 0.0
        delta += 0.15 if grape  else 0.0
        delta += 0.20 if garlic else 0.0
        delta += drug_eff
        delta -= missed * 0.35
        delta += {'None':0.0,'Mild':0.15,'Moderate':0.35,'Severe':0.65}[illness]
        delta += 0.30 if diarhea else 0.0

        inr = round(float(np.clip(base + delta + np.random.normal(0, 0.20), 0.8, 10.0)), 1)

        if   inr < 2.0: status = 'Subtherapeutic'
        elif inr <= 3.5: status = 'Therapeutic'
        elif inr <= 5.0: status = 'Supratherapeutic'
        else:            status = 'Critical High'

        # ── Dose adjustment ────────────────────────────────────────────────
        old_dose = dose
        if   inr < 1.5:  pct, action =  0.20, 'Increase'
        elif inr < 2.0:  pct, action =  0.10, 'Increase'
        elif inr < 2.5:  pct, action =  0.05, 'Increase'
        elif inr <= 3.5: pct, action =  0.00, 'Maintain'
        elif inr <= 4.0: pct, action = -0.05, 'Decrease'
        elif inr <= 5.0: pct, action = -0.15, 'Decrease'
        else:            pct, action = -0.25, 'Hold/Decrease'

        new_dose = round(np.clip(dose * (1 + pct), 0.5, 20.0) * 2) / 2

        # ── Adverse events ─────────────────────────────────────────────────
        bleed_p = 0.005 if inr <= 4.0 else (0.025 if inr <= 5.0 else 0.07)
        bleed   = random.choices(['None','Minor','Major'], weights=[1-bleed_p-0.001, bleed_p, 0.001])[0]
        clot_p  = 0.01 if inr >= 2.5 else (0.03 if inr >= 2.0 else 0.06)
        clot    = 'Yes' if random.random() < clot_p else 'No'
        hosp    = 'Yes' if (bleed == 'Major' or clot == 'Yes') else 'No'

        # ── Clinical notes ─────────────────────────────────────────────────
        notes = []
        if inr > 5.0:          notes.append('Critical INR – consider Vitamin K reversal')
        if inr < 1.5:          notes.append('Severely subtherapeutic – high stroke/thrombosis risk')
        if drug_eff >  0.8:    notes.append(f'Significant potentiation by {drug_name}')
        if drug_eff < -0.8:    notes.append(f'Significant INR reduction by {drug_name} – increase dose')
        if missed >= 2:        notes.append(f'{missed} missed doses reported this week')
        if vitk == 'High':     notes.append('High Vitamin K intake – may reduce INR')

        weekly_records.append({
            'Patient_ID': pt['Patient_ID'],
            'Week_Number': wk,
            'Visit_Date': visit_dt.strftime('%Y-%m-%d'),
            'INR_Result': inr,
            'INR_Status': status,
            'Current_Warfarin_Dose_mg_day': old_dose,
            'Dose_Action': action,
            'New_Warfarin_Dose_mg_day': new_dose,
            'VitK_Intake_Level': vitk,
            'VitK_Foods_Consumed': vitk_f,
            'Alcohol_Units_Per_Week': alc,
            'Cranberry_Juice': 'Yes' if cranb  else 'No',
            'Grapefruit':      'Yes' if grape  else 'No',
            'Garlic_Supplement':'Yes' if garlic else 'No',
            'Concurrent_Medication': drug_name,
            'Medication_Category': drug_cat,
            'Medication_INR_Effect': 'Increase' if drug_eff>0 else ('Decrease' if drug_eff<0 else 'None'),
            'Missed_Doses_This_Week': missed,
            'Illness_Event': illness,
            'Diarrhea_Vomiting': 'Yes' if diarhea else 'No',
            'Exercise_Level': exercise,
            'Bleeding_Event': bleed,
            'Thromboembolic_Event': clot,
            'Hospitalization': hosp,
            'Clinical_Notes': '; '.join(notes) if notes else ''
        })
        dose = new_dose

weekly_df = pd.DataFrame(weekly_records)
print(f"Weekly records generated: {len(weekly_df)}")

# ── Drug Interaction Reference ────────────────────────────────────────────────
drug_ref = pd.DataFrame([
    ('Amoxicillin',       'Antibiotic',            'Increase', 'Mild (+0.3–0.6)',          'Gut flora reduction → ↓ VitK synthesis',           'Monitor INR; may need small dose reduction'),
    ('Ciprofloxacin',     'Antibiotic (FQ)',        'Increase', 'Moderate (+0.6–1.2)',      'CYP1A2 inhibition + gut flora reduction',           'Check INR 3–5 days after starting; reduce warfarin 15–25%'),
    ('Metronidazole',     'Antibiotic',             'Increase', 'Strong (+1.0–2.0)',        'CYP2C9 inhibition',                                  'Reduce warfarin 25–50%; monitor INR closely'),
    ('Trimethoprim-SMX',  'Antibiotic',             'Increase', 'Strong (+0.6–1.5)',        'CYP2C9 inhibition + VitK reduction',                'Reduce warfarin; check INR within a week'),
    ('Fluconazole',       'Antifungal (Azole)',     'Increase', 'Strong (+1.5–3.0)',        'Potent CYP2C9 & CYP3A4 inhibitor',                  'Reduce warfarin 30–50%; check INR in 2–3 days'),
    ('Itraconazole',      'Antifungal (Azole)',     'Increase', 'Moderate (+1.0–2.0)',      'CYP3A4 inhibition',                                  'Reduce warfarin 20–40%; monitor INR'),
    ('Amiodarone',        'Antiarrhythmic',         'Increase', 'Moderate-Strong (+0.8–1.5)','CYP2C9 & CYP3A4 inhibition; very long half-life', 'Reduce warfarin 30–50%; effect persists months after stopping'),
    ('Ibuprofen/NSAIDs',  'NSAID',                 'Increase', 'Mild (+0.2–0.5)',          'COX-1 inhibition ↑ bleeding risk; mild CYP2C9',     'Avoid if possible; if used, monitor INR & GI bleeding'),
    ('Omeprazole',        'PPI',                   'Increase', 'Mild (+0.1–0.3)',          'Weak CYP2C19 inhibition',                            'Low risk; routine monitoring sufficient'),
    ('Simvastatin',       'Statin',                'Increase', 'Mild (+0.1–0.25)',         'Mild CYP3A4 competition',                            'Monitor INR when starting or changing statin dose'),
    ('Levothyroxine',     'Thyroid Hormone',       'Increase', 'Variable (+0.3–1.0)',      '↑ Catabolism of clotting factors',                   'Adjust warfarin when thyroid status changes'),
    ('Rifampin',          'Antibiotic (Rifamycin)','Decrease', 'Strong (−1.0 to −2.5)',   'Potent CYP2C9 & CYP3A4 inducer',                    'May need to double warfarin; monitor very closely; effect reverses on discontinuation'),
    ('Carbamazepine',     'Anticonvulsant',        'Decrease', 'Moderate-Strong (−1.0–1.5)','CYP enzyme induction',                             'Increase warfarin; monitor INR carefully'),
    ('Aspirin 81 mg',     'Antiplatelet',          'Mild Inc', 'Mild but ↑ bleeding risk', 'Platelet inhibition synergizes with anticoagulation','Use lowest effective dose; monitor for bleeding'),
],
columns=['Drug_Name','Category','Effect_on_INR','Magnitude','Mechanism','Clinical_Action'])

# ── Food Interaction Reference ─────────────────────────────────────────────────
food_ref = pd.DataFrame([
    ('Spinach (cooked)',  494, 'Decrease',      'High',         'Highest VitK; consistency of intake is key'),
    ('Kale (cooked)',     817, 'Decrease',      'Very High',    'Extreme VitK; even moderate intake can significantly lower INR'),
    ('Collard Greens',    623, 'Decrease',      'Very High',    'Monitor INR if dietary intake changes'),
    ('Broccoli',          141, 'Decrease',      'Moderate',     'Large quantities may lower INR'),
    ('Brussels Sprouts',  177, 'Decrease',      'Moderate-High','Maintain consistent intake'),
    ('Cabbage',            76, 'Decrease',      'Moderate',     'Manageable with consistent dosing'),
    ('Lettuce (raw)',     102, 'Decrease',      'Moderate',     'Large daily salads can affect INR'),
    ('Cranberry Juice',     5, 'Increase',      'Moderate',     'CYP2C9 inhibition – can raise INR despite low VitK'),
    ('Grapefruit',          0, 'Increase',      'Mild-Moderate','CYP3A4 inhibitor; effect on warfarin is mild but monitor'),
    ('Alcohol >14 u/wk',    0, 'Increase',      'Moderate-High','Inhibits warfarin metabolism; binge drinking can cause dangerous elevation'),
    ('Garlic supplement',  17, 'Increase',      'Mild',         'Mild antiplatelet effect; high-dose supplements ↑ bleeding risk'),
    ('Avocado',            21, 'Decrease (large)','Low-Moderate','Contains VitK; may reduce warfarin absorption in large amounts'),
    ('Mango (large qty)',   4, 'Increase',      'Low',          'Anecdotal CYP inhibition; monitor with large daily intake'),
    ('Soy milk/products',  29, 'Decrease',      'Low-Moderate', 'Contains VitK; consistent moderate intake is manageable'),
],
columns=['Food','VitK_per_100g_mcg','INR_Effect','Impact_Level','Notes'])

# ── Summary statistics per patient ────────────────────────────────────────────
summary = weekly_df.groupby('Patient_ID').agg(
    Mean_INR                    = ('INR_Result',           'mean'),
    Min_INR                     = ('INR_Result',           'min'),
    Max_INR                     = ('INR_Result',           'max'),
    INR_StdDev                  = ('INR_Result',           'std'),
    Pct_Time_Therapeutic        = ('INR_Status',           lambda x: round((x=='Therapeutic').mean()*100,1)),
    Pct_Time_Subtherapeutic     = ('INR_Status',           lambda x: round((x=='Subtherapeutic').mean()*100,1)),
    Pct_Time_Supratherapeutic   = ('INR_Status',           lambda x: round(((x=='Supratherapeutic')|(x=='Critical High')).mean()*100,1)),
    Total_Dose_Adjustments      = ('Dose_Action',          lambda x: (x != 'Maintain').sum()),
    Total_Missed_Doses          = ('Missed_Doses_This_Week','sum'),
    Weeks_With_Drug_Interaction = ('Medication_INR_Effect',lambda x: (x != 'None').sum()),
    Bleeding_Events             = ('Bleeding_Event',       lambda x: (x != 'None').sum()),
    Thromboembolic_Events       = ('Thromboembolic_Event', lambda x: (x=='Yes').sum()),
    Hospitalizations            = ('Hospitalization',      lambda x: (x=='Yes').sum()),
).reset_index().round(2)

# ── Excel formatting helpers ──────────────────────────────────────────────────
def write_sheet(ws, df, title, hdr_color='1F4E79'):
    if title:
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(name='Arial', bold=True, size=13)
        ws.row_dimensions[1].height = 22
        data_row_start = 3
    else:
        data_row_start = 1

    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=data_row_start, column=ci, value=col.replace('_',' '))
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill      = PatternFill('solid', start_color=hdr_color)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[data_row_start].height = 28

    alt_fill = PatternFill('solid', start_color='EBF3FB')
    for ri, row in df.iterrows():
        xr = data_row_start + ri + 1
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=xr, column=ci, value=val)
            c.font      = Font(name='Arial', size=9)
            c.alignment = Alignment(vertical='center')
            if ri % 2 == 0:
                c.fill = alt_fill

    # Column widths
    for ci, col in enumerate(cols, 1):
        sample_vals = [str(df.iloc[i, ci-1]) for i in range(min(30, len(df)))]
        max_len = max(len(col), max((len(v) for v in sample_vals), default=8))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 45)

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = Workbook()

ws1 = wb.active
ws1.title = 'Patient Demographics'
write_sheet(ws1, patients_df.drop(columns=['_sens']),
            'Mechanical Mitral Valve Cohort — Patient Demographics (N=250)',
            hdr_color='1F4E79')

ws2 = wb.create_sheet('Weekly INR Records')
write_sheet(ws2, weekly_df,
            f'Warfarin Management — Weekly Longitudinal Records  ({N_WEEKS} weeks × {N_PATIENTS} patients = {len(weekly_df):,} rows)',
            hdr_color='1A5276')

ws3 = wb.create_sheet('Drug Interactions Reference')
write_sheet(ws3, drug_ref,
            'Drug–Warfarin Interaction Reference Table',
            hdr_color='6C3483')

ws4 = wb.create_sheet('Food Interactions Reference')
write_sheet(ws4, food_ref,
            'Food–Warfarin Interaction Reference Table',
            hdr_color='1E8449')

ws5 = wb.create_sheet('Patient Summary Statistics')
write_sheet(ws5, summary,
            'Per-Patient Summary Statistics (20-Week Observation Period)',
            hdr_color='784212')

out_path = '/sessions/focused-determined-mayer/mnt/outputs/warfarin_cohort_MMV.xlsx'
wb.save(out_path)
print(f"\nFile saved: {out_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"\nDataset summary:")
print(f"  Patients:           {len(patients_df)}")
print(f"  Weekly records:     {len(weekly_df):,}")
drug_interaction_weeks = (weekly_df['Medication_INR_Effect'] != 'None').sum()
print(f"  Drug interaction wks: {drug_interaction_weeks:,}")
bleeding_count = (weekly_df['Bleeding_Event'] != 'None').sum()
print(f"  Bleeding events:    {bleeding_count}")
print(f"  Thromboembolic evts:{(weekly_df['Thromboembolic_Event']=='Yes').sum()}")
print(f"  Mean INR (cohort):  {weekly_df['INR_Result'].mean():.2f}")
print(f"  % Time therapeutic: {(weekly_df['INR_Status']=='Therapeutic').mean()*100:.1f}%")
